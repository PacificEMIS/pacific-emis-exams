# ---
# jupyter:
#   jupytext:
#     formats: ipynb,py:percent
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.3'
#       jupytext_version: 1.16.2
#   kernelspec:
#     display_name: Python 3 (ipykernel)
#     language: python
#     name: python3
# ---

# %%
# %%time
###############################################################################
# This notebook focuses on processing data from excel spreadsheet directly    #
# into another format ready to load into OnlineSBA                            #
#                                                                             #
# It is also used to process raw SOE workbooks and flag data issues           #
###############################################################################

# import everything we need throughout the notebook
# core stuff
import itertools
import os
from pathlib import Path
import json
import re

# Data stuff
import pandas as pd # Data analysis
import xlrd # excel 
from openpyxl import Workbook # excel
from sqlalchemy.engine import URL # SQL DB
import numpy as np

# Fuzzy searching stuff
from fuzzywuzzy import fuzz
# process is used to compare a string to MULTIPLE other strings
from fuzzywuzzy import process

# Pretty printing stuff
from tqdm.notebook import trange, tqdm
from IPython.display import display, HTML
import pprint
pp = pprint.PrettyPrinter(indent=4)

# Generate unique identifiers stuff
import uuid
import random

rd = random.Random()
rd.seed(0)

# Configuration (initial setup)
with open('config.json', 'r') as file:
     config = json.load(file)

test = config['test']
country = config['country']
cwd = os.getcwd()

year_to_load = config['load_year']
skip_incorrect_answers = config['skip_incorrect_answers']
flag_duplicate_students = config['flag_duplicate_students']
remove_items_metadata = config['remove_items_metadata']
export = config['export']
fix_schoolid_in_source_data = config['fix_schoolid_in_source_data']
accept_teachers_with_three_chars_only = config['accept_teachers_with_three_chars_only']
accept_unknown_gender = config['accept_unknown_gender']
accept_unknown_student = config['accept_unknown_student']
accept_unknown_teacher = config['accept_unknown_teacher']

# Establish a database server connection
conn = """
    Driver={{ODBC Driver 17 for SQL Server}};
    Server={},{};
    Database={};
    authentication=SqlPassword;UID={};PWD={};
    TrustServerCertificate=yes;
    autocommit=True
    """.format(config['server_ip'], config['server_port'], config['database'], config['uid'], config['pwd'])

connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conn})

from sqlalchemy import create_engine
engine = create_engine(connection_url)

# It is important to keep the order of the cells since there are inplace 
# operations on DataFrames

# %%
# Load the schools, student enrollments and teachers from the database
# For students and teachers currenly only those of that year of loaded and used to compare with exams data
# e.g. student enrolled in 2018-19 and teachers teaching in 2018-19 and compared with exams data
# for 2018-19
import pandas as pd
import sqlalchemy as sa

query_student_enrol = """
SELECT
	stuCardID
	, CONCAT(stuGiven,' ',stuFamilyName) AS Student -- stuMiddleNames,' ',
	, stuGender
	, stuDoB
	, schNo
	, stueYear
	FROM Student_ S
	INNER JOIN StudentEnrolment_ SE ON S.stuID = SE.stuID
"""

query_schools = """
SELECT
	schNo
	, schName
	FROM Schools
"""

# Not used yet
#query_teachers = """
#"""

with engine.begin() as sql_conn:
    #df = pd.read_sql_query(sa.text("SELECT 'thing' as txt"), sql_conn)
    
    df_student_enrol = pd.read_sql_query(query_student_enrol, sql_conn)
    print('df_student_enrol')
    display(df_student_enrol.head(3))

    df_schools = pd.read_sql_query(query_schools, sql_conn)
    print('df_schools')
    display(df_schools.head(3))


# %%
def load_excel_to_df(filename):
    """Loads an Excel filename to a Pandas DataFrame.

    Parameters
    ----------
    filename : str, required
        The filename of the excel file to load

    Raises
    ------
    NotImplementedError
        Could raise unknown error. Implement if it happens
    
    Returns
    -------
    DataFrame
    """
    file_path = Path(filename)
    file_extension = file_path.suffix.lower()[1:]

    if file_extension == 'xlsx':
        df_student_results = pd.read_excel(filename, index_col=None, header=0, engine='openpyxl')
    elif file_extension == 'xls':
        df_student_results = pd.read_excel(filename, index_col=None, header=0)
    elif file_extension == 'csv':
        df_student_results = pd.read_csv(filename, index_col=None, header=0)
    else:
        raise Exception("File not supported")

    return df_student_results


# %%
# Load a single SOE Assessment workbook (for testing,)
# in particular the sheet with the raw data
local_path = os.path.abspath('/mnt/h/Development/Pacific EMIS/repositories-data/pacific-emis-exams/')
#filename = os.path.join(local_path, 'RMI/MISAT/MISAT 2019/3GrEng2019/AllSchools_A03_2018-19_Results.xls')
#filename = os.path.join(local_path, 'RMI/MISAT/MISAT 2012/6grEng12/AllSchools_A06_2011-12_Results.xls')
#filename = os.path.join(local_path, 'RMI/MISAT/MISAT 2019/Gr6Math2019/AllSchools_M06_2018-19_Results.xls')
#filename = os.path.join(local_path, 'RMI/MISAT/MISAT 2009/6GrMath2009/AllSchools_M06_2008-09_Results.xls')
#filename = os.path.join(local_path, 'RMI/MISAT/MISAT 2009/Gr6KM2009/AllSchools_B06_2008-09_Results.xls')
#filename = os.path.join(local_path, 'RMI/MISAT/MISAT 2009/3GrMath2009/AllSchools_M03_2008-09_Results.xls')
#filename = os.path.join(local_path, 'RMI/MISAT/MISAT 2011/3GrEng2011/AllSchools_A03_2010-11_Results1.xls')
#filename = os.path.join(local_path, 'RMI/MISAT/MISAT 2018/Gr3KM2018/AllSchools_B03_2017-18_Results.xls')
#filename = os.path.join(local_path, 'RMI/MISAT/MISAT 2016/Gr8HSET2016/AllSchools_H08_2015-16_Results.xls')
#filename = os.path.join(local_path, 'RMI/MISAT/MISAT 2017/Gr8HSET2017/AllSchools_H08_2016-17_Results.xls')
#filename = os.path.join(local_path, 'RMI/MISAT/MISAT 2009/8GrHSET2009/AllSchools_H08_2008-09_Results.xls')
#filename = os.path.join(local_path, 'RMI/MISAT/MISAT 2017/Gr3Math2017/AllSchools_M03_2016-17_Results.xls')
#filename = os.path.join(local_path, 'RMI/MISAT/MISAT 2017/Gr10Math2017/AllSchools_M10_2016-17_Results1.xls')
filename = os.path.join(local_path, 'RMI/MISAT/MISAT 2018/Gr10Math2018/AllSchools_M10_2017-18_Results.xls')
#filename = os.path.join(local_path, 'FSM/NMCT/NMCT 2021/AllSchools_R08_2020-21_Results.xls')

testname = filename.split('/')[-1]
df_student_results = {}
df_student_results[testname] = load_excel_to_df(filename)
print('df_student_results')
display(df_student_results[testname])

# %%
# %%time
# Load all SOE Assessment workbook inside a directory
# (~32 seconds on iMac with i9 CPU and 32GB RAM)
cwd = os.getcwd()
path = os.path.join(local_path,country+'/'+test)

if year_to_load != 'all':
    path = os.path.join(path, year_to_load)

df_student_results_list = {}

for root, directories, files in os.walk(path, topdown=False):
    for name in files:
        filename = os.path.join(root, name)
        print('Loading into DataFrame file:', filename)
        try:
            df_student_results_list[name] = load_excel_to_df(filename)
        except:
            print('Problem loading file:', filename)
            #print('Error was:', )            

print('Completed loading excel files')


# %%
def merge_exams_data_with_student_enrol_df(df_student_results, df_student_enrol, testing=False):
    """ Merge both the dirty exams data with the clean student enrollments dataset

    Parameters
    ----------
    df_student_results : DataFrame, required
        The student results DataFrame (from SOE Assessment response sheet)
    df_student_enrol : DataFrame, required
        The student enrolment DataFrame (from EMIS)
        
    Raises
    ------
    NotImplementedError
        Could raise unknown error. Implement if it happens
    
    Returns
    -------
    DataFrame
    """
    
    # lower case to make join case insensitive (like SQL Server, the default collation of Pacific EMIS anyway)
    try:
        df_student_results['StudentName2'] = df_student_results['StudentName'].str.lower()
        df_student_enrol['Student2'] = df_student_enrol['Student'].str.lower()
    except KeyError:        
        print('StudentName column is not present or misspelled (hint from data): ', df_student_results[:1].iloc[:, : 5].to_csv(index=False, header=False))
        return
    except:
        print('Unknown error')
        return

    # Also need to trim spaces to make it exactly like the SQL Server join
    df_student_results['StudentName2'] = df_student_results['StudentName2'].str.strip()
    df_student_enrol['Student2'] = df_student_enrol['Student2'].str.strip()

    # Before we attempt to merge
    # Only keep one of the duplicates from the EMIS
    df_student_enrol.drop_duplicates(keep='last', inplace=True)
    if testing: print('Total student enrol: ', len(df_student_enrol.index))

    # isolate into a seperate DataFrame students with
    # same name but different DoB, school, etc. (i.e. different students of same name)
    df_student_enrol.duplicated(subset=['Student2'])
    df_student_enrol_nonambiguous = df_student_enrol[~df_student_enrol.duplicated(subset=['Student2'], keep=False)]
    df_student_enrol_ambiguous = df_student_enrol[df_student_enrol.duplicated(subset=['Student2'], keep=False)]
    if testing: print('Total student enrol that are not ambiguous: ', len(df_student_enrol_nonambiguous.index))
    if testing: print('Total student enrol that are ambiguous: ', len(df_student_enrol_ambiguous.index))
    if testing: print('Check ambiguous + not ambiguous equals all enrolled (minus duplicates): ', len(df_student_enrol_nonambiguous.index) + len(df_student_enrol_ambiguous.index))
    if testing: 
        print('df_student_enrol_nonambiguous') 
        display(df_student_enrol_nonambiguous.head(2))
    df_student_enrol_ambiguous.sort_values(by=['Student2'])

    # For now, process using only non-ambiguous student enrolment records
    # It would only be possible to use non-ambiguous student enrolment records
    # if the exams data would contain the correct school, DoB or other data
    # that could disambiguate students with same name

    # Merge student exams data with student enrolments
    df_students_results_and_enrol = df_student_results.set_index('StudentName2').join(df_student_enrol_nonambiguous.set_index('Student2'), lsuffix='_caller', rsuffix='_other')
    df_students_results_and_enrol = df_student_results.merge(df_student_enrol_nonambiguous, how='left', left_on='StudentName2', right_on='Student2', suffixes=('_from_exams', '_from_db'), indicator=False)
    if testing: 
        print('df_students_results_and_enrol') 
        display(df_students_results_and_enrol.head(2))
    
    return df_students_results_and_enrol


# %%
# Merge student exams data with student enrollments
# Working with the single student exams file (for testing)
df_students_results_and_enrol = {}
df_students_results_and_enrol[testname] = merge_exams_data_with_student_enrol_df(df_student_results[testname], df_student_enrol, True)
print('df_students_results_and_enrol')
df_students_results_and_enrol[testname]

# %%
# %%time
# Merge student exams data with student enrollments
# Working with all student exams files (~22 seconds on iMac with i9 CPU and 32GB RAM)
df_students_results_and_enrol_list = {}

for file,df in tqdm(df_student_results_list.items()):
    df_students_results_and_enrol_list[file] = merge_exams_data_with_student_enrol_df(df, df_student_enrol, False)

df_students_results_and_enrol_list
# Remove any None item from list (those DataFrames could not be merged)
#df_students_results_and_enrol_list = list(filter(lambda x: x is not None, df_students_results_and_enrol_list))
for k in tqdm(df_students_results_and_enrol_list):
    if df_students_results_and_enrol_list[k] is None:
        del df_students_results_and_enrol_list[k]
        tqdm.write("None DataFrame, could not be merged, investigate file {}".format(k))

# %%
df_students_results_and_enrol_list[list(df_students_results_and_enrol_list.keys()).pop()]

# %%
# This list is to be confirmed and updated as necessary
# If a school name is in an exam file but not in here we need to generate an error message
# and update this list with the correct mapping to the canonical school ID
schools_lookup_from_exams_byname = {
    ' Ebeye Christian -Private': 'KWA105',
    ' Ebeye SDA-Private Primary': 'KWA109',
    ' Ine-Arno': 'ARN103',
    'Aerok A-Aelonlaplap': 'AIL100',       
    'Aerok A-Ailinglaplap': 'AIL100',
    'Aerok A-Maloelap': 'MAL101',
    'Aerok A-Maloeplap': 'MAL101',
    'Aerok A-Medium': 'AIL100',
    'Aerok A-Public ': 'AIL100',
    'Aerok A-Public': 'AIL100',
    'Aerok M-Maleolap': 'MAL101',
    'Aerok M-Maloelap': 'MAL101',
    'Aerok M-Northern': 'MAL101',
    'Aerok M-Public': 'MAL101',
    'Aerok Protestant-Private': 'AIL109', 
    'Aerok, A-Ailinglaplap': 'AIL100',
    'Ailuk -Ailuk ': 'ALU101',
    'Ailuk -Ailuk': 'ALU101',
    'Ailuk Protestant-Private': 'ALU103',
    'Ailuk -Public ': 'ALU101',
    'Ailuk -Public': 'ALU101',
    'Ailuk-Ailuk': 'ALU101',
    'Ailuk-Ailuk': 'ALU101',
    'Ailuk-Enejelaar': 'ALU102',
    'Ailuk-Medium': 'ALU101',
    'Ailuk-Northern': 'ALU101',
    'Ailuk-Public ': 'ALU101',
    'Ailuk-Public': 'ALU101',
    'Airok A-Aelonlaplap': 'AIL100',
    'Airok A-Ailinglaplap': 'AIL100',
    'Airok A-Central': 'AIL100',
    'Airok A-Public ': 'AIL100',
    'Airok A-Public': 'AIL100',
    'Airok M-Maloelap': 'MAL101',
    'Airok M-Maloeplap': 'MAL101',
    'Airok M-Mejit': 'MAL101',
    'Airok M-Public': 'MAL101',
    'Airok M-Small': 'MAL101',
    'Airok Protestant-Private': 'AIL109',
    'Airok  M-Maloelap': 'MAL101',
    'Airok  M-Public': 'MAL101',
    'Airok, M-Maloelap': 'MAL101',
    'Ajeltake Chistian Academy-Majuro': 'MAJ102',
    'Ajeltake Chistian Academy-Private Primary': 'MAJ102',
    'Ajeltake Christian Academy-Majuro': 'MAJ102',
    'Ajeltake Christian Academy-Private ': 'MAJ102',
    'Ajeltake Christian Academy-Private Primary': 'MAJ102',
    'Ajeltake Christian Academy-Private': 'MAJ102',
    'Ajeltake Christian Academy-Public ': 'MAJ102',
    'Ajeltake Christian Acedemy-Majuro': 'MAJ102',
    'Ajeltake Christian Acedemy-Private': 'MAJ102',
    'Ajeltake -Majuro': 'MAJ101',
    'Ajeltake-Large': 'MAJ101',
    'Ajeltake-Majuro': 'MAJ101',
    'Ajeltake-Majuro': 'MAJ101',
    'Ajeltake-Public ': 'MAJ101',
    'Ajeltake-Public': 'MAJ101',
    'Arno -Arno ': 'ARN101',
    'Arno -Arno': 'ARN101',
    'Arno -Public ': 'ARN101',
    'Arno -Public': 'ARN101',
    'Arno-Arno': 'ARN101',
    'Arno-Arno': 'ARN101',
    'Arno-Eastern': 'ARN101',
    'Arno-Medium': 'ARN101',
    'Arno-Public ': 'ARN101',
    'Arno-Public': 'ARN101',
    'Assumption High -Private Secondary': 'MAJ104',
    'Assumption High School-Public Secondary': 'MAJ104',
    'Assumption HS-Ailinglaplap': 'MAJ104',
    'Assumption -Private': 'MAJ103',
    'Assumption-Majuro': 'MAJ103',
    'Assumption-Private ': 'MAJ103',
    'Assumption-Private Primary': 'MAJ103',
    'Assumption-Private Secondary': 'MAJ104',
    'Assumption-Private': 'MAJ103',
    'Assumption-Private': 'MAJ103',
    'Assumption-Public ': 'MAJ103',
    'Aur -Aur ': 'AUR101',
    'Aur -Aur': 'AUR101',
    'Aur -Public ': 'AUR101',
    'Aur -Public': 'AUR101',
    'Aur-Aur': 'AUR101',
    'Aur-Aur': 'AUR101',
    'Aur-Medium': 'AUR101',
    'Aur-Northern': 'AUR101',
    'Aur-Public ': 'AUR101',
    'Aur-Public': 'AUR101',
    'Bikarej -Arno': 'ARN102',
    'Bikarej -Public ': 'ARN102',
    'Bikarej-Arno ': 'ARN102',
    'Bikarej-Arno': 'ARN102',
    'Bikarej-Arno': 'ARN102',
    'Bikarej-Eastern': 'ARN102',
    'Bikarej-Medium': 'ARN102',
    'Bikarej-Public ': 'ARN102',
    'Bikarej-Public': 'ARN102',
    'Bouj -Aelonlaplap': 'AIL101',
    'Bouj -Ailinglaplap': 'AIL101',
    'Bouj -Public': 'AIL101',
    'Bouj-Aelonlaplap': 'AIL101',
    'Bouj-Ailinglaplap': 'AIL101',
    'Bouj-Public': 'AIL101',
    'Buoj -Ailinglaplap': 'AIL101',
    'Buoj-Aelonlaplap': 'AIL101',
    'Buoj-Ailinglaplap': 'AIL101',
    'Buoj-Central': 'AIL101',
    'Buoj-Medium': 'AIL101',
    'Buoj-Public ': 'AIL101',
    'Buoj-Public': 'AIL101',
    'Carlos -Kwajalein': 'KWA101',
    'Carlos -Public ': 'KWA101',
    'Carlos-Kwajalein': 'KWA101',
    'Carlos-Kwajalein': 'KWA101',
    'Carlos-Kwajlein': 'KWA101',
    'Carlos-Public ': 'KWA101',
    'Carlos-Public': 'KWA101',
    'Carlos-Small': 'KWA101',
    'Deaf Center-Majuro': 'MAJ131',
    'Delap Calvary-Private': 'KSA103',
    'Delap -Majuro': 'MAJ105',
    'Delap SDA High -Private Secondary': 'MAJ108',
    'Delap SDA High School-Public Secondary': 'MAJ108',
    'Delap SDA HS-Ailinglaplap': 'MAJ108',
    'Delap SDA -Majuro': 'MAJ107',
    'Delap SDA -Private': 'MAJ107',
    'Delap SDA-Majuro': 'MAJ107',
    'Delap SDA-Private ': 'MAJ107',
    'Delap SDA-Private Primary': 'MAJ107',
    'Delap SDA-Private Secondary': 'MAJ108',
    'Delap SDA-Private': 'MAJ107',
    'Delap SDA-Private': 'MAJ107',
    'Delap SDA-Public ': 'MAJ107',
    'Delap  -Majuro': 'MAJ105',
    'Delap  -Public': 'MAJ105',
    'Delap-Majuro': 'MAJ105',
    'Delap-Public': 'MAJ105',
    'DES-Large': 'MAJ105',
    'DES-Majuro': 'MAJ105',
    'DES-Majuro': 'MAJ105',
    'DES-Public ': 'MAJ105',
    'DES-Public': 'MAJ105',
    'Ebadon -Kwajalein': 'KWA102',
    'Ebadon -Public ': 'KWA102',
    'Ebadon -Public': 'KWA102',
    'Ebadon-Kwajalein': 'KWA102',
    'Ebadon-Kwajalein': 'KWA102',
    'Ebadon-Kwajlein': 'KWA102',
    'Ebadon-Public ': 'KWA102',
    'Ebadon-Public': 'KWA102',
    'Ebadon-Small': 'KWA102',
    'Ebeye Calvary High -Private Secondary': 'KWA104',
    'Ebeye Calvary High School-Private Secondary': 'KWA104',
    'Ebeye Calvary HS-Ailinglaplap': 'KWA104',
    'Ebeye Calvary -Kwajalein': 'KWA103',
    'Ebeye Calvary -Private': 'KWA103',
    'Ebeye Calvary-Private Primary': 'KWA103',
    'Ebeye Calvary-Private Secondary': 'KWA104',
    'Ebeye Calvary-Private': 'KWA103',
    'Ebeye Calvary-Private': 'KWA103',
    'Ebeye Calvary-Private': 'KWA103',
    'Ebeye Calvary-Public ': 'KWA103',
    'Ebeye Calvary-Public': 'KWA103',
    'Ebeye Cavalry-Private Primary': 'KWA103',
    'Ebeye Christian-Kwajalein': 'KWA105',
    'Ebeye Christian-Private ': 'KWA105',
    'Ebeye Christian-Private Primary': 'KWA105',
    'Ebeye Christian-Private': 'KWA105',
    'Ebeye Christian-Private': 'KWA105',
    'Ebeye Christian-Private': 'KWA105',
    'Ebeye Christian-Public ': 'KWA105',
    'Ebeye Deaf Center -Kwajalein': 'KWA121',
    'Ebeye Deaf Center -Kwajelein': 'KWA121',
    'Ebeye Deaf Center -Private Secondary': 'KWA120',
    'Ebeye Deaf Edu. -Private Secondary': 'KWA120',
    'Ebeye Elementary-Kwajelein': 'KWA108',
    'Ebeye Middle Public-Kwajalein': 'KWA107',
    'Ebeye Middle School-Kwajalein': 'KWA107',
    'Ebeye Middle School-Public': 'KWA107',
    'Ebeye Public -Public ': 'KWA108',
    'Ebeye Public-Kwajalein': 'KWA108',
    'Ebeye Public-Kwajelein': 'KWA108',
    'Ebeye Public-Kwajlein': 'KWA108',
    'Ebeye Public-Large': 'KWA108',
    'Ebeye Public-Public ': 'KWA108',
    'Ebeye Public-Public': 'KWA108',
    'Ebeye SDA High -Private Secondary': 'KWA110',
    'Ebeye SDA High School-Private Secondary': 'KWA110',
    'Ebeye SDA HS-Ailinglaplap': 'KWA110',
    'Ebeye SDA -Kwajalein': 'KWA109',
    'Ebeye SDA -Private': 'KWA109',
    'Ebeye SDA-Private ': 'KWA109',
    'Ebeye SDA-Private Primary': 'KWA109',
    'Ebeye SDA-Private Secondary': 'KWA110',
    'Ebeye SDA-Private': 'KWA109',
    'Ebeye SDA-Private': 'KWA109',
    'Ebeye SDA-Public ': 'KWA109',
    'Ebon -Ebon ': 'EBO101',
    'Ebon -Ebon': 'EBO101',
    'Ebon -Public ': 'EBO101',
    'Ebon -Public': 'EBO101',
    'Ebon-Ebon': 'EBO101',
    'Ebon-Ebon': 'EBO101',
    'Ebon-Medium': 'EBO101',
    'Ebon-Public ': 'EBO101',
    'Ebon-Public': 'EBO101',
    'Ebon-Southern': 'EBO101',
    'EES/ Ejit??-Kili': 'KIL101',
    'Ejit-Kili ': 'KIL101',
    'Ejit-Kili/Bikini': 'KIL101',
    'Ejit-Kili': 'KIL101',
    'Ejit-Majuro': 'KIL101',
    'Ejit-Medium': 'KIL101',
    'Ejit-Public ': 'KIL101',
    'Ejit-Public': 'KIL101',
    'Ejit-Southern': 'KIL101',
    'Enburr-Kwajalein': 'KWA111',
    'Enejelaar -Ailuk ': 'ALU102',
    'Enejelaar -Ailuk': 'ALU102',
    'Enejelaar-Ailuk': 'ALU102',
    'Enejelaar-Alluk': 'ALU102',
    'Enejelaar-Northern': 'ALU102',
    'Enejelaar-Public ': 'ALU102',
    'Enejelaar-Public': 'ALU102',
    'Enejelaar-Small': 'ALU102',
    'Enejet -Mili': 'MIL101',
    'Enejet -Public ': 'MIL101',
    'Enejet -Public': 'MIL101',
    'Enejet-Eastern': 'MIL101',
    'Enejet-Enejet': 'MIL101',
    'Enejet-Medium': 'MIL101',
    'Enejet-Mili': 'MIL101',
    'Enejet-Public ': 'MIL101',
    'Enejet-Public': 'MIL101',
    'Enekoion -Ebon': 'EBO102',
    'Enekoion -Public': 'EBO102',
    'Enekoion-Ebon': 'EBO102',
    'Enekoion-Ebon': 'EBO102',
    'Enekoion-Public ': 'EBO102',
    'Enekoion-Public': 'EBO102',
    'Enekoion-Small': 'EBO102',
    'Enekoion-Southern': 'EBO102',
    'Enewa -Aelonlaplap': 'AIL102',
    'Enewa -Ailinglaplap': 'AIL102',
    'Enewa -Public ': 'AIL102',
    'Enewa -Public': 'AIL102',
    'Enewa-Aelonlaplap': 'AIL102',
    'Enewa-Ailinglaplap': 'AIL102',
    'Enewa-Central': 'AIL102',
    'Enewa-Public ': 'AIL102',
    'Enewa-Public': 'AIL102',
    'Enewa-Small': 'AIL102',
    'Enewetak -Enewetak ': 'ENE101',
    'Enewetak -Public ': 'ENE101',
    'Enewetak-Eastern': 'ENE101',
    'Enewetak-Enewetak': 'ENE101',
    'Enewetak-Enewetak': 'ENE101',
    'Enewetak-Public ': 'ENE101',
    'Enewetak-Public': 'ENE101',
    'Enniburr High School-Public Secondary': 'KWA119',
    'Enniburr -Kwajalein': 'KWA111',
    'Enniburr -Public ': 'KWA111',
    'Enniburr -Public': 'KWA111',
    'Enniburr-Kwajalein': 'KWA111',
    'Enniburr-Kwajalein': 'KWA111',
    'Enniburr-Kwajlein': 'KWA111',
    'Enniburr-Medium': 'KWA111',
    'Enniburr-Public ': 'KWA111',
    'Enniburr-Public Secondary': 'KWA119',
    'Enniburr-Public': 'KWA111',
    'Father Hacker High School-Private Secondary': 'KWA118',
    'Father Hacker HS-Ailinglaplap': 'KWA118',
    'Father Hacker-Private Secondary': 'KWA118',
    'Gem Chirstian School-Private Primary': 'KWA112',
    'Gem Christian High School-Private Secondary': 'KWA113',
    'Gem Christian -Private ': 'KWA112',
    'Gem Christian School-Kwajalein': 'KWA113',
    'Gem Christian School-Private ': 'KWA112',
    'Gem Christian School-Private Primary': 'KWA112',
    'Gem Christian School-Private': 'KWA112',
    'Gem Christian School-Private': 'KWA112',
    'Gem Christian School-Private': 'KWA112',
    'Gem Christian School-Public ': 'KWA112',
    'Gem High School-Private Secondary': 'KWA113',
    'Gem HS-Ailinglaplap': 'KWA113',
    'Gem -Private': 'KWA112',
    'Gem-Private Secondary': 'KWA113',
    'Imiej -Jaluit ': 'JAL101',
    'Imiej -Jaluit': 'JAL101',
    'Imiej -Public ': 'JAL101',
    'Imiej -Public': 'JAL101',
    'Imiej-Jaluit': 'JAL101',
    'Imiej-Jaluit': 'JAL101',
    'Imiej-Medium': 'JAL101',
    'Imiej-Public ': 'JAL101',
    'Imiej-Public': 'JAL101',
    'Imiej-Southern': 'JAL101',
    'Imroj -Jaluit': 'JAL102',
    'Imroj Protestant-Private': 'JAL110',
    'Imroj -Public ': 'JAL102',
    'Imroj -Public': 'JAL102',
    'Imroj -Southern': 'JAL102',
    'Imroj-Jaluit': 'JAL102',
    'Imroj-Jaluit': 'JAL102',
    'Imroj-Medium': 'JAL102',
    'Imroj-Public ': 'JAL102',
    'Imroj-Public': 'JAL102',
    'Imroj-Southern': 'JAL102',
    'Ine -Arno': 'ARN103',
    'Ine -Public ': 'ARN103',
    'Ine -Public': 'ARN103',
    'Ine-Arno ': 'ARN103',
    'Ine-Arno': 'ARN103',
    'Ine-Arno': 'ARN103',
    'Ine-Eastern': 'ARN103',
    'Ine-Medium': 'ARN103',
    'Ine-Public ': 'ARN103',
    'Ine-Public': 'ARN103',
    'Jabat -Jabat': 'JAB101',
    'Jabat-Central': 'JAB101',
    'Jabat-Jabat': 'JAB101',
    'Jabat-Public ': 'JAB101',
    'Jabat-Public': 'JAB101',
    'Jabnoden -Jaluit ': 'JAL103',
    'Jabnoden -Jaluit': 'JAL103',
    'Jabnoden -Public ': 'JAL103',
    'Jabnoden -Public': 'JAL103',
    'Jabnoden-Jaluit': 'JAL103',
    'Jabnoden-Jaluit': 'JAL103',
    'Jabnoden-Public ': 'JAL103',
    'Jabnoden-Public': 'JAL103',
    'Jabnodren-Jaluit': 'JAL103',
    'Jabnodren-Public ': 'JAL103',
    'Jabnodren-Southern': 'JAL103',
    'Jabonden-Jaluit': 'JAL103',
    'Jabor -Jaluit ': 'JAL104',
    'Jabor -Jaluit': 'JAL104',
    'Jabor -Public ': 'JAL104',
    'Jabor -Public': 'JAL104',
    'Jabor-Jaluit': 'JAL104',
    'Jabor-Jaluit': 'JAL104',
    'Jabor-Medium': 'JAL104',
    'Jabor-Public ': 'JAL104',
    'Jabor-Public': 'JAL104',
    'Jabor-Southern': 'JAL104',
    'Jabot -Public': 'JAB101',
    'Jabro -Private': 'KWA115',
    'Jabro-Private': 'KWA115',
    'Jah -Ailinglaplap': 'AIL103',
    'Jah -Public ': 'AIL103',
    'Jah -Public': 'AIL103',
    'Jah-Aelonlaplap': 'AIL103',
    'Jah-Ailinglaplap': 'AIL103',
    'Jah-Central': 'AIL103',
    'Jah-Public ': 'AIL103',
    'Jah-Public': 'AIL103',
    'Jah-Small': 'AIL103',
    'Jaluit -Jaluit ': 'JAL105',
    'Jaluit -Jaluit': 'JAL105',
    'Jaluit -Public ': 'JAL105',
    'Jaluit -Public': 'JAL105',
    'Jaluit-???': 'JAL105',
    'Jaluit-Jaljuit': 'JAL105',
    'Jaluit-Jaluit': 'JAL105',
    'Jaluit-Jaluit': 'JAL105',
    'Jaluit-Medium': 'JAL105',
    'Jaluit-Public ': 'JAL105',
    'Jaluit-Public': 'JAL105',
    'Jaluit-Southern': 'JAL105',
    'Jang -Public ': 'MAL102',
    'Jang-Maloelap': 'MAL102',
    'Jang-Maloelap': 'MAL102',
    'Jang-Maloeplap': 'MAL102',
    'Jang-Northern': 'MAL102',
    'Jang-Public ': 'MAL102',
    'Jang-Public': 'MAL102',
    'Jang-Small': 'MAL102',
    'Japo -Arno': 'ARN104',
    'Japo -Public ': 'ARN104',
    'Japo-Arno ': 'ARN104',
    'Japo-Arno': 'ARN104',
    'Japo-Arno': 'ARN104',
    'Japo-Eastern': 'ARN104',
    'Japo-Medium': 'ARN104',
    'Japo-Public ': 'ARN104',
    'Japo-Public': 'ARN104',
    'Jebal -Likiep ': 'LIK101',
    'Jebal -Likiep': 'LIK101',
    'Jebat-Jebat': 'JAB101',
    'Jebro High School-Private Secondary': 'KWA114',
    'Jebro High School-Public Secondary': 'KWA114',
    'Jebro HS-Ailinglaplap': 'KWA114',
    'Jebro Kabua-Private': 'KWA115',
    'Jebro-Kwajalein': 'KWA115',
    'Jebro-Private ': 'KWA115',
    'Jebro-Private Primary': 'KWA115',
    'Jebro-Private Secondary': 'KWA114',
    'Jebro-Private': 'KWA115',
    'Jebro-Private': 'KWA115',
    'Jebro-Private': 'KWA115',
    'Jebro-Public ': 'KWA115',
    'Jebwan -Ailinglaplap': 'AIL105',
    'Jebwan-Aelonlaplap': 'AIL105',
    'Jebwan-Ailinglaplap': 'AIL105',
    'Jebwan-Central': 'AIL105',
    'Jebwan-Public ': 'AIL105',
    'Jebwan-Public': 'AIL105',
    'Jebwan-Small': 'AIL105',
    'Jeh -Ailinglaplap': 'AIL104',
    'Jeh -Public ': 'AIL104',
    'Jeh -Public': 'AIL104',
    'Jeh SDA-Private': 'AIL110',
    'Jeh-Aelonlaplap': 'AIL104',
    'Jeh-Ailinglaplap': 'AIL104',
    'Jeh-Central': 'AIL104',
    'Jeh-Medium': 'AIL104',
    'Jeh-Public ': 'AIL104',
    'Jeh-Public': 'AIL104',
    'Jepal -Public ': 'LIK101',
    'Jepal-Likiep': 'LIK101',
    'Jepal-Likiep': 'LIK101',
    'Jepal-Northern': 'LIK101',
    'Jepal-Public ': 'LIK101',
    'Jepal-Public': 'LIK101',
    'Jepal-Small': 'LIK101',
    'JHS-Ailinglaplap': 'JAL106',
    'JHS-Public Secondary': 'JAL106',
    'Jobwon -Ailinglaplap': 'AIL105',
    'KAHS-Ailinglaplap': 'KWA116',
    'KAHS-Public Secondary': 'KWA116',
    'Kattiej -Aelonlaplap': 'AIL106',
    'Kattiej -Ailinglaplap': 'AIL106',
    'Kattiej -Public ': 'AIL106',
    'Kattiej-Aelonlaplap': 'AIL106',
    'Kattiej-Ailinglaplap': 'AIL106',
    'Kattiej-Central': 'AIL106',
    'Kattiej-Public ': 'AIL106',
    'Kattiej-Public': 'AIL106',
    'Kattiej-Small': 'AIL106',
    'Kaven -Maloelap': 'MAL103',
    'Kaven -Public ': 'MAL103',
    'Kaven-Maleolap': 'MAL103',
    'Kaven-Maloelap': 'MAL103',
    'Kaven-Maloeplap': 'MAL103',
    'Kaven-Northern': 'MAL103',
    'Kaven-Public ': 'MAL103',
    'Kaven-Public': 'MAL103',
    'Kaven-Small': 'MAL103',
    'Kilange -Arno': 'ARN105',
    'Kilange -Public ': 'ARN105',
    'Kilange -Public': 'ARN105',
    'Kilange-Arno ': 'ARN105',
    'Kilange-Arno': 'ARN105',
    'Kilange-Arno': 'ARN105',
    'Kilange-Eastern': 'ARN105',
    'Kilange-Medium': 'ARN105',
    'Kilange-Public ': 'ARN105',
    'Kilange-Public': 'ARN105',
    'Kili -Kili ': 'KIL102',
    'Kili -Kili': 'KIL102',
    'Kili -Public ': 'KIL102',
    'Kili -Southern': 'KIL102',
    'Kili-Kili/Bikini': 'KIL102',
    'Kili-Kili': 'KIL102',
    'Kili-Medium': 'KIL102',
    'Kili-Public ': 'KIL102',
    'Kili-Public': 'KIL102',
    'Kili-Southern': 'KIL102',
    'Kinange-Arno': 'ARN105',
    'Lae -Lae ': 'LAE101',
    'Lae -Lae': 'LAE101',
    'Lae -Public ': 'LAE101',
    'Lae -Public': 'LAE101',
    'Lae-Lae ': 'LAE101',
    'Lae-Lae': 'LAE101',
    'Lae-Medium': 'LAE101',
    'Lae-Public ': 'LAE101',
    'Lae-Public': 'LAE101',
    'Lae-Western': 'LAE101',
    'Laura Christian Academy-Majuro': 'MAJ129',
    'Laura Christian Academy-Private': 'MAJ129',
    'Laura High School-Majuro': 'MAJ111',
    'Laura High School-Public': 'MAJ111',
    'Laura -Majuro': 'MAJ109',
    'Laura Protestant-Private': 'MAJ133',
    'Laura Public -Public ': 'MAJ109',
    'Laura Public-Large': 'MAJ109',
    'Laura Public-Majuro': 'MAJ109',
    'Laura Public-Majuro': 'MAJ109',
    'Laura Public-Public ': 'MAJ109',
    'Laura Public-Public': 'MAJ109',
    'Laura SDA -Majuro': 'MAJ110',
    'Laura SDA -Private': 'MAJ110',
    'Laura SDA-Majuro': 'MAJ110',
    'Laura SDA-Private ': 'MAJ110',
    'Laura SDA-Private Primary': 'MAJ110',
    'Laura SDA-Private': 'MAJ110',
    'Laura SDA-Public ': 'MAJ110',
    'Laura-Majuro': 'MAJ109',
    'Laura-Public ': 'MAJ109',
    'LHS -Ailinglaplap': 'MAJ111',
    'LHS-Ailinglaplap': 'MAJ111',
    'LHS-Majuro': 'MAJ111',
    'LHS-Public Secondary': 'MAJ111',
    'LHS-Public': 'MAJ111',
    'Lib -Lib ': 'LIB101',
    'Lib -Lib': 'LIB101',
    'Lib-Lib': 'LIB101',
    'LIB-LIB': 'LIB101',
    'Lib-Medium': 'LIB101',
    'Lib-Public ': 'LIB101',
    'Lib-Public': 'LIB101',
    'Lib-Western': 'LIB101',
    'Life Skills Academy-Majuro': 'MAJ113',
    'Life Skills Academy-Public': 'MAJ113',
    'Likiep -Likiep ': 'LIK102',
    'Likiep -Likiep': 'LIK102',
    'Likiep -Public ': 'LIK102',
    'Likiep -Public': 'LIK102',
    'Likiep-Likiep': 'LIK102',
    'Likiep-Likiep': 'LIK102',
    'Likiep-Medium': 'LIK102',
    'Likiep-Northern': 'LIK102',
    'Likiep-Public ': 'LIK102',
    'Likiep-Public': 'LIK102',
    'Loen -Namu ': 'NAU101',
    'Loen -Namu': 'NAU101',
    'Loen -Public ': 'NAU101',
    'Loen -Public': 'NAU101',
    'Loen-Central': 'NAU101',
    'Loen-Medium': 'NAU101',
    'Loen-Namu': 'NAU101',
    'Loen-Namu': 'NAU101',
    'Loen-Public ': 'NAU101',
    'Loen-Public': 'NAU101',
    'Long Isand-Majuro': 'MAJ112',
    'Long Island-Majuro': 'MAJ112',
    'Long Island-Public ': 'MAJ112',
    'Long Island-Public': 'MAJ112',
    'Longar -Arno': 'ARN106',
    'Longar -Public ': 'ARN106',
    'Longar-Arno ': 'ARN106',
    'Longar-Arno': 'ARN106',
    'Longar-Arno': 'ARN106',
    'Longar-Eastern': 'ARN106',
    'Longar-Medium': 'ARN106',
    'Longar-Public ': 'ARN106',
    'Longar-Public': 'ARN106',
    'Lukoj-Arno': 'ARN107',
    'Lukoj-Eastern': 'ARN107',
    'Lukoj-Public': 'ARN107',
    'Lukoj-Small': 'ARN107',
    'Lukonwod -Mili': 'MIL102',
    'Lukonwod -Public': 'MIL102',
    'Lukonwod-Eastern': 'MIL102',
    'Lukonwod-Enewetak': 'ENE101',
    'Lukonwod-Lukonwod': 'MIL102',
    'Lukonwod-Mili': 'MIL102',
    'Lukonwod-Mili': 'MIL102',
    'Lukonwod-Public ': 'MIL102',
    'Lukonwod-Public': 'MIL102',
    'Lukonwod-Small': 'MIL102',
    'Lukunwod-Mili': 'MIL102',
    'Mae -Namu ': 'NAU102',
    'Mae -Public ': 'NAU102',
    'Mae-Central': 'NAU102',
    'Mae-Namu': 'NAU102',
    'Mae-Namu': 'NAU102',
    'Mae-Public ': 'NAU102',
    'Mae-Public': 'NAU102',
    'Majken -Namu': 'NAU103',
    'Majken-Central': 'NAU103',
    'Majken-Medium': 'NAU103',
    'Majken-Namu': 'NAU103',
    'Majken-Public ': 'NAU103',
    'Majken-Public': 'NAU103',
    'Majkin -Namu': 'NAU103',
    'Majkin -Public': 'NAU103',
    'Majkin-Central': 'NAU103',
    'Majkin-Namu': 'NAU103',
    'Majkin-Public ': 'NAU103',
    'Majkin-Public': 'NAU103',
    'Majuro Baptist Academy-Private Primary': 'MAJ114',
    'Majuro Baptist Christian Academy-Majuro': 'MAJ114',
    'Majuro Baptist Christian Academy-Private ': 'MAJ114',
    'Majuro Baptist Christian Academy-Private Primary': 'MAJ114',
    'Majuro Baptist Christian Academy-Private Secondary': 'MAJ115',
    'Majuro Baptist Christian Academy-Private': 'MAJ114',
    'Majuro Baptist Christian Academy-Private': 'MAJ114',
    'Majuro Baptist Christian Academy-Public ': 'MAJ114',
    'Majuro Baptist Christian-Private': 'MAJ114',
    'Majuro Baptist HS-Ailinglaplap': 'MAJ115',
    'Majuro Baptist-Private': 'MAJ114',
    'Majuro Coop High -Private Secondary': 'MAJ117',
    'Majuro Coop HS-Ailinglaplap': 'MAJ117',
    'Majuro Cooperative High School-Private Secondary': 'MAJ117',
    'Majuro Coop-Majuro': 'MAJ116',
    'Majuro Coop-Private ': 'MAJ116',
    'Majuro Coop-Private Primary': 'MAJ116',
    'Majuro Coop-Private Secondary': 'MAJ117',
    'Majuro Coop-Private': 'MAJ116',
    'Majuro Coop-Private': 'MAJ116',
    'Majuro Coop-Public ': 'MAJ116',
    'Majuro Deaf Center -Public ': 'MAJ131',
    'Majuro Deaf Center-Majuro': 'MAJ131',
    'Majuro Deaf Center-Private Secondary': 'MAJ132',
    'Majuro Deaf School-Majuro': 'MAJ131',
    'Majuro Middle School-Majuro': 'MAJ120',
    'Majuro Middle School-Public': 'MAJ120',
    'Marshall Christian High School-Private Secondary': 'MAJ118',
    'Marshall Christian High-Private Secondary': 'MAJ118',
    'Marshall Christian-Private Secondary': 'MAJ118',
    'Marshall Islands High School-Majuro': 'MAJ118',
    'Marshall Islands High School-Public': 'MAJ118',
    'Marshalls Christian High School-Private': 'MAJ118',
    'Matolen -Arno': 'ARN108',
    'Matolen -Public': 'ARN108',
    'Matolen-Arno ': 'ARN108',
    'Matolen-Arno': 'ARN108',
    'Matolen-Arno': 'ARN108',
    'Matolen-Eastern': 'ARN108',
    'Matolen-Medium': 'ARN108',
    'Matolen-Public ': 'ARN108',
    'Matolen-Public': 'ARN108',
    'MCHS -Ailinglaplap': 'MAJ117',
    'MCHS-Private Secondary': 'MAJ118',
    'MCHS-Public Secondary': 'MAJ118',
    'MDEC-Majuro': 'MAJ131',
    'MDEC-Public': 'MAJ131',
    'MDED-Majuro': 'MAJ131',
    'Mejatto -Mejatto ': 'RON101',
    'Mejatto -Public': 'RON101',
    'Mejatto -Rongelap': 'RON101',
    'Mejatto-Kwajalein': 'RON101',
    'Mejatto-Maloeplap': 'RON101',
    'Mejatto-Medium': 'RON101',
    'Mejatto-Mejatto': 'RON101',
    'Mejatto-Public': 'RON101',
    'Mejatto-Western': 'RON101',
    'Mejel -Ailinglaplap': 'AIL107',
    'Mejel -Public': 'AIL107',
    'Mejel  -Aelonlaplap': 'AIL107',
    'Mejel-Aelonlaplap': 'AIL107',
    'Mejel-Ailinglaplap': 'AIL107',
    'Mejel-Central': 'AIL107',
    'Mejel-Public ': 'AIL107',
    'Mejel-Public': 'AIL107',
    'Mejel-Small': 'AIL107',
    'Mejirirok -Jaluit': 'JAL107',
    'Mejit -Mejit ': 'MEJ101',
    'Mejit -Mejit': 'MEJ101',
    'Mejit -Public ': 'MEJ101',
    'Mejit -Public': 'MEJ101',
    'Mejit-Medium': 'MEJ101',
    'Mejit-Mejit': 'MEJ101',
    'Mejit-Mejit': 'MEJ101',
    'Mejit-Northern': 'MEJ101',
    'Mejit-Public ': 'MEJ101',
    'Mejit-Public': 'MEJ101',
    'Mejrirok -Jaluit ': 'JAL107',
    'Mejrirok -Jaluit': 'JAL107',
    'Mejrirok-Jaluit': 'JAL107',
    'Mejrirok-Public': 'JAL107',
    'Mejrirok-Southern': 'JAL107',
    'Mejurirok-Jaluit': 'JAL107',
    'Mejurirok-Medium': 'JAL107',
    'Mejurirok-Public ': 'JAL107',
    'Mejurirok-Public': 'JAL107',
    'Mejurirok-Southern': 'JAL107',
    'Melang -Likiep': 'LIK103',
    'Melang -Public': 'LIK103',
    'Melang-Likiep': 'LIK103',
    'Melang-Northern': 'LIK103',
    'Melang-Public ': 'LIK103',
    'Melang-Public': 'LIK103',
    'Melang-Small': 'LIK103',
    'Melan-Likiep': 'LIK103',
    'Melan-Public ': 'LIK103',
    'Melan-Public': 'LIK103',
    'Middle School-Public': 'MAJ120',
    'MIHS -Ailinglaplap': 'MAJ119',
    'MIHS-Public Secondary': 'MAJ119',
    'Mili -Mili': 'MIL103',
    'Mili -Public': 'MIL103',
    'Mili-Eastern': 'MIL103',
    'Mili-Mili': 'MIL103',
    'Mili-Mili': 'MIL103',
    'Mili-Public ': 'MIL103',
    'Mili-Public': 'MIL103',
    'MMS-Majuro': 'MAJ120',
    'Nallo -Mili': 'MIL104',
    'Nallo -Public ': 'MIL104',
    'Nallo -Public': 'MIL104',
    'Nallo-Eastern': 'MIL104',
    'Nallo-Jaluit': 'MIL104',
    'Nallo-Medium': 'MIL104',
    'Nallo-Mili': 'MIL104',
    'Nallo-Mili': 'MIL104',
    'Nallo-Public ': 'MIL104',
    'Nallo-Public': 'MIL104',
    'Namdrik -Namdrik ': 'NAM101',
    'Namdrik -Namdrik': 'NAM101',
    'Namdrik -Public ': 'NAM101',
    'Namdrik -Public': 'NAM101',
    'Namdrik-Large': 'NAM101',
    'Namdrik-Namdrik': 'NAM101',
    'Namdrik-Namdrik': 'NAM101',
    'Namdrik-Public ': 'NAM101',
    'Namdrik-Public': 'NAM101',
    'Namdrik-Southern': 'NAM101',
    'Namu -Namu ': 'NAU104',
    'Namu-Central': 'NAU104',
    'Namu-Namu': 'NAU104',
    'Namu-Public ': 'NAU104',
    'Namu-Public': 'NAU104',
    'Namu-Small': 'NAU104',
    'Narmej -Jaluit': 'JAL108',
    'Narmej -Public ': 'JAL108',
    'Narmej-Jaluit': 'JAL108',
    'Narmej-Jaluit': 'JAL108',
    'Narmej-Medium': 'JAL108',
    'Narmej-Public ': 'JAL108',
    'Narmej-Public': 'JAL108',
    'Narmej-Southern': 'JAL108',
    'Narmij-Jaluit ': 'JAL108',
    'Narmij-Southern': 'JAL108',
    'NDES -Majuro': 'MAJ126',
    'NDES-Majuro': 'MAJ126',
    'NIHS -Ailinglaplap': 'WTH101',
    'NIHS -Public Secondary': 'WTH101',
    'NIHS-Private Secondary': 'WTH101',
    'NIHS-Public Secondary': 'WTH101',
    'NIHS-Wotje': 'WTH101',
    'North Delap-Majuro': 'MAJ126',
    'North Delap-Public ': 'MAJ126',
    'North Delap-Public': 'MAJ126',
    'NVTI-Majuro': 'MAJ128',
    'Ollet -Maloelap': 'MAL104',
    'Ollet -Maloeplap': 'MAL104',
    'Ollet -Ollet ': 'MAL104',
    'Ollet -Public ': 'MAL104',
    'Ollet -Public': 'MAL104',
    'Ollet-Maloelap': 'MAL104',
    'Ollet-Maloelap': 'MAL104',
    'Ollet-Maloeplap': 'MAL104',
    'Ollet-Northern': 'MAL104',
    'Ollet-Public ': 'MAL104',
    'Ollet-Public': 'MAL104',
    'Ollet-Small': 'MAL104',
    'Queen of Peace-Kwajalein': 'KWA117',
    'Queen of Peace-Private ': 'KWA117',
    'Queen of Peace-Private Primary': 'KWA117',
    'Queen of Peace-Private': 'KWA117',
    'Queen of Peace-Private': 'KWA117',
    'Queen of Peace-Private': 'KWA117',
    'Queen of Peace-Public ': 'KWA117',
    'Rairok -Majuro': 'MAJ121',
    'Rairok -Public ': 'MAJ121',
    'Rairok -Public': 'MAJ121',
    'Rairok-Large': 'MAJ121',
    'Rairok-Majuro ': 'MAJ121',
    'Rairok-Majuro': 'MAJ121',
    'Rairok-Majuro': 'MAJ121',
    'Rairok-Public ': 'MAJ121',
    'Rairok-Public': 'MAJ121',
    'RES-Large': 'MAJ122',
    'RES-Majuro': 'MAJ122',
    'RES-Majuro': 'MAJ122',
    'RES-Public ': 'MAJ122',
    'RES-Public': 'MAJ122',
    'Rita Chrisitan-Private': 'MAJ123',
    'Rita Christian High School-Private Secondary': 'MAJ124',
    'Rita Christian High-Private Secondary': 'MAJ124',
    'Rita Christian HS-Ailinglaplap': 'MAJ124',
    'Rita Christian -Private': 'MAJ123',
    'Rita Christian-Majuro': 'MAJ123',
    'Rita Christian-nan': 'MAJ124',
    'Rita Christian-Private ': 'MAJ123',
    'Rita Christian-Private Primary': 'MAJ123',
    'Rita Christian-Private Secondary': 'MAJ124',
    'Rita Christian-Private Seconday': 'MAJ124',
    'Rita Christian-Private': 'MAJ123',
    'Rita Christian-Private': 'MAJ123',
    'Rita Christian-Public ': 'MAJ123',
    'Rita -Majuro': 'MAJ122',  
    'Rita-Majuro': 'MAJ122',
    'Rita-Public': 'MAJ122',
    'Rongrong Christian Academy-Private': 'MAJ125',
    'Rongrong Christian Elementary-Private': 'MAJ125',
    'Rongrong Christian High-Private Secondary': 'MAJ118',
    'Rongrong Christian School-Private': 'MAJ118',
    'RongRong Christian-Private Primary': 'MAJ125',
    'Rongrong Christian-Private': 'MAJ125',
    'Rongrong Elementary-Majuro': 'MAJ125',
    'Rongrong Elementary-Private': 'MAJ125',
    'RongROng -Majuro': 'MAJ125',
    'Rongrong-Majuro': 'MAJ125',
    'RongRong-Majuro': 'MAJ125',
    'Rongrong-Private ': 'MAJ125',
    'Rongrong-Private': 'MAJ125',
    'RongRong-Private': 'MAJ125',
    'Rongrong-Public ': 'MAJ125',
    'Rongrong-RongRong': 'MAJ125',
    'Rongrong-Rongrong': 'MAJ125',
    'Ronrong-Ronrong': 'MAJ125',
    'St. Joseph -Jaluit': 'JAL109',
    'St. Joseph -Private': 'JAL109',
    'St. Joseph-Jaluit': 'JAL109',
    'St. Joseph-Private ': 'JAL109',
    'St. Joseph-Private Primary': 'JAL109',
    'St. Joseph-Private': 'JAL109',
    'St. Joseph-Private': 'JAL109',
    'St. Joseph-Private': 'JAL109',
    'St. Paul-Arno': 'ARN109',
    'St. Paul-Private': 'ARN109',
    'St. Paul-Private': 'ARN109',
    'St. Thomas-Private': 'WTH104', 
    'St. Thomas-Wotje': 'WTH104',
    'St.Joseph-Jaluit': 'JAL109',
    'St.Joseph-Private': 'JAL109',
    'St.Joseph-Public ': 'JAL109',
    'St.Paul-Arno': 'ARN109',
    'Tarawa -Maloelap': 'MAL105',
    'Tarawa -Public ': 'MAL105',
    'Tarawa -Public': 'MAL105',
    'Tarawa-Majuro': 'MAL105',
    'Tarawa-Maloelap': 'MAL105',
    'Tarawa-Maloelap': 'MAL105',
    'Tarawa-Maloeplap': 'MAL105',
    'Tarawa-Medium': 'MAJ105',
    'Tarawa-Northern': 'MAL105',
    'Tarawa-Public ': 'MAL105',
    'Tarawa-Public': 'MAL105',
    'Tinak -Arno': 'ARN109',
    'Tinak -Public ': 'ARN109',
    'Tinak -Public': 'ARN109',
    'Tinak-Arno ': 'ARN109',
    'Tinak-Arno': 'ARN109',
    'Tinak-Eastern': 'ARN109',
    'Tinak-Medium': 'ARN109',
    'Tinak-Public ': 'ARN109',
    'Tinak-Public': 'ARN109',
    'Tobal -Aur': 'AUR102',
    'Tobal -Public ': 'AUR102',
    'Tobal-Aur': 'AUR102',
    'Tobal-Aur': 'AUR102',
    'Tobal-Medium': 'AUR102',
    'Tobal-Northern': 'AUR102',
    'Tobal-Public ': 'AUR102',
    'Tobal-Public': 'AUR102',
    'Toka -Ebon': 'EBO103',
    'Toka -Public ': 'EBO103',
    'Toka -Public': 'EBO103',
    'Toka-Ebon': 'EBO103',
    'Toka-Ebon': 'EBO103',
    'Toka-Medium': 'EBO103',
    'Toka-Public ': 'EBO103',
    'Toka-Public': 'EBO103',
    'Toka-Southern': 'EBO103',
    'Tokewa -Mili': 'MIL105',
    'Tokewa -Public ': 'MIL105',
    'Tokewa -Public': 'MIL105',
    'Tokewa-Eastern': 'MIL105',
    'Tokewa-Mili': 'MIL105',
    'Tokewa-Public ': 'MIL105',
    'Tokewa-Public': 'MIL105',
    'Tokewa-Small': 'MIL105',
    'Tutu -Arno': 'ARN110',
    'Tutu -Public ': 'ARN110',
    'Tutu-Arno ': 'ARN110',
    'Tutu-Arno': 'ARN110',
    'Tutu-Eastern': 'ARN110',
    'Tutu-Public ': 'ARN110',
    'Tutu-Public': 'ARN110',
    'Tutu-Small': 'ARN110',
    'UES-Large': 'ARN111',
    'UES-Majuro': 'ARN111',
    'UES-Majuro': 'MAJ126',
    'UES-Public ': 'ARN111',
    'UES-Public': 'ARN111',
    'Ujae -Public ': 'UJA101',
    'Ujae -Public': 'UJA101',
    'Ujae -Ujae ': 'UJA101',
    'Ujae -Ujae': 'UJA101',
    'Ujae-Medium': 'UJA101',
    'Ujae-Public ': 'UJA101',
    'Ujae-Public': 'UJA101',
    'Ujae-Ujae ': 'UJA101',
    'Ujae-Ujae': 'UJA101',
    'Ujae-Western': 'UJA101',
    'Ulien -Arno': 'ARN111',
    'Ulien -Public ': 'ARN111',
    'Ulien -Public': 'ARN111',
    'Ulien-Arno ': 'ARN111',
    'Ulien-Arno': 'ARN111',
    'Ulien-Arno': 'ARN111',
    'Ulien-Eastern': 'ARN111',
    'Ulien-Medium': 'ARN111',
    'Ulien-Public ': 'ARN111',
    'Ulien-Public': 'ARN111',
    'Uliga -Majuro': 'MAJ126',
    'Uliga Protestant-Private': 'MAJ130',
    'Uliga-Majuro': 'MAJ126',
    'UPCS-Majuro': 'MAJ130',
    'UPCS-Private': 'MAJ130',
    'Utrik -Public ': 'UTR101',
    'Utrik -Utrik ': 'UTR101',
    'Utrik-Medium': 'UTR101',
    'Utrik-Northern': 'UTR101',
    'Utrik-Public ': 'UTR101',
    'Utrik-Public': 'UTR101',
    'Utrik-Utrik': 'UTR101',
    'Utrik-Utrik': 'UTR101',
    'Wodmeej -Wotje': 'WTH102',
    'Wodmeej-Northern': 'WTH102',
    'Wodmeej-Public ': 'WTH102',
    'Wodmeej-Public': 'WTH102',
    'Wodmeej-Small': 'WTH102',
    'Wodmeej-Wotje ': 'WTH102',
    'Wodmeej-Wotje': 'WTH102',
    'Wodmeej-Wotje': 'WTH102',
    'Wodmej -Public ': 'WTH102',
    'Wodmej -Wotje': 'WTH102',
    'Wodmej-Public': 'WTH102',
    'Woja A -Aelonlaplap': 'AIL108',
    'Woja A -Ailinglaplap': 'AIL108',
    'Woja A -Public ': 'AIL108',
    'Woja A-Aelonlaplap': 'AIL108',
    'Woja A-Ailinglaplap': 'AIL108',
    'Woja A-Central': 'AIL108',
    'Woja A-Medium': 'AIL108',
    'Woja A-Public ': 'AIL108',
    'Woja A-Public': 'AIL108',
    'Woja M -Majuro': 'MAJ127',
    'Woja M -Public ': 'MAJ127',
    'Woja M-Large': 'MAJ127',
    'Woja M-Majuro': 'MAJ127',
    'Woja M-Majuro': 'MAJ127',
    'Woja M-Private': 'MAJ127',
    'Woja M-Public ': 'MAJ127',
    'Woja M-Public': 'MAJ127',
    'Woja SDA-Private': 'AIL111',
    'Woja, A-Ailinglaplap': 'AIL108',
    'Wotho-Wotho': 'WOT101',
    'Wotje -Wotje ': 'WTH103',
    'Wotje -Wotje': 'WTH103',
    'Wotje-Large': 'WTH103',
    'Wotje-Northern': 'WTH103',
    'Wotje-Public ': 'WTH103',
    'Wotje-Public': 'WTH103',
    'Wotje-Wotje': 'WTH103',
    'Wotje-Wotje': 'WTH103',
    'Wotto -Wotto ': 'WOT101',
    'Wotto-Public ': 'WOT101',
    'Wotto-Public': 'WOT101',
    'Wotto-Small': 'WOT101',
    'Wotto-Western': 'WOT101',
    'Wotto-Wotto': 'WOT101',
    'Narmij-Jaluit': 'JAL108',
    'Jang-Majuro': 'MAL102',
    'Kaven-Majuro': 'MAL103',
}


# %%
def clean_exams(df, name, testing=False):
    """ Does any cleanup/validation needed with Exams ID/Name.

    Parameters
    ----------    
    df: DataFrame, required
        The student results and enrol DataFrame
    name: str, required
        The name of the excel file this DataFrame came from
    testing: bool, required
        Whether we are test (usually single DataFrame)    
        
    Raises
    ------
    NotImplementedError
        Could raise unknown error. Implement if it happens
    
    Returns
    -------
    DataFrame
    """
    expected_testid = name.split('_')[1]
    
    # Basic check of the TestID in the filename
    p = re.compile('[A-Z]{1}[0-9]{2}')
    if not p.match(expected_testid):
        print("The TestID is not what was expected. Make sure the excel file {} follows the format AllSchools_TESTID_YYYY-YY_Results.xls".format(name))
    
    # Check the TestID in the data is has expected (i.e. what is in the filename)
    testid_fromdata = df['TestID'].iloc[0]
    if testid_fromdata != expected_testid:
        print("The TestID in the data is not the same as the expected test ID from the excel filename for {}".format(name))
        
    # Check that all the TestID are the same in the data
    if len(df['TestID'].unique()) != 1:
        print("The TestID in the data is not unique and contains unexpected values (e.g. A03 mixed with B03) from the excel filename for {}".format(name))
        
    # No actual cleanup at the moment, just flag data validation issues
    if testing:
        print('Cleaned exams for file {}'.format(name))
        #display(df)
        
    return df
    
def clean_schools(df, df_schools, name, testing=False):
    """ Does any cleanup/validation needed with SchoolIDs.

    Parameters
    ----------    
    df: DataFrame, required
        The student results and enrol DataFrame
    df_schools : DataFrame, required
        The schools DataFrame (from EMIS)
    name: str, required
        The name of the excel file this DataFrame came from
    testing: bool, required
        Whether we are test (usually single DataFrame) 
        
    Raises
    ------
    NotImplementedError
        Could raise unknown error. Implement if it happens
    
    Returns
    -------
    DataFrame
    """        

    # From EMIS, get school ID to name official mapping
    schools_lookup = df_schools.set_index('schNo').to_dict()['schName']
    schools_lookup_byname = df_schools.set_index('schName').to_dict()['schNo']
    
    # If the Results or Data file contains correct SchoolID can ignore the name mapping
    # and just use that to produce the correct SchoolName
    # All valid school IDs in the EMIS
    school_ids = list(df_schools['schNo'])
    
    if testing: 
        print('schools_lookup')
        #pp.pprint(dict(schools_lookup))
        pp.pprint(dict(itertools.islice(schools_lookup.items(), 3)))
        
        print('schools_lookup_byname')
        #pp.pprint(dict(schools_lookup_byname))
        pp.pprint(dict(itertools.islice(schools_lookup_byname.items(), 3)))
    
    schools_lookup_from_exams = {y:x for x,y in schools_lookup_from_exams_byname.items()}
    
    if testing:
        print('schools_lookup_from_exams_byname')
        #pp.pprint(dict(schools_lookup_from_exams_byname))
        pp.pprint(dict(itertools.islice(schools_lookup_from_exams_byname.items(), 3)))

    # ??? Check if this is primary or elementary, some have same school names so use
    # grade of test to define the school

    # Create a temporary SchoolName and SchoolIsland joined
    df['SchoolNameTemp'] = df.agg('{0[SchoolName]}-{0[IslandName]}'.format, axis=1)
    if testing:
        print('Cleaning schools SchoolNameTemp')
        display(df['SchoolNameTemp'])
    
    # Upper case all school ID
    # Convert to string, Not needed perhaps?
    df['SchoolID'] = df['SchoolID'].astype(str)
    df['SchoolID'] = df['SchoolID'].str.upper()
    # Strip spaces
    df['SchoolID'] = df['SchoolID'].str.strip()
    
    # Check if the school ID in the exams data file exists in the EMIS
    # and create temporary school name for those
    #s_school_ids1 = df['SchoolID'].map(schools_lookup_byname)
    #df = df.assign(SchoolIDTemp1 = s_school_ids1)
    mask = df['SchoolID'].isin(df_schools['schNo'].values)
    df['SchoolIDTemp1'] = df['SchoolID'].where(mask)
    # Check if the school name in the exams data file has a mapping hard coded (old/incorrect schoolIDs)
    # and create temporary school name for those
    s_school_ids2 = df['SchoolNameTemp'].map(schools_lookup_from_exams_byname)
    df = df.assign(SchoolIDTemp2 = s_school_ids2)

    if testing:
        print('Cleaning schools SchoolIDTemp1')
        display(df['SchoolIDTemp1'])
        print('Cleaning schools SchoolIDTemp2')
        display(df['SchoolIDTemp2'])
        
    # Coalesce to get the school ID
    # Use bfill if I end up using more then two columns to coalesce
    # https://stackoverflow.com/questions/38152389/coalesce-values-from-2-columns-into-a-single-column-in-a-pandas-dataframe
    df['SchoolIDFinal'] = df.SchoolIDTemp1.combine_first(df.SchoolIDTemp2)
    
    # An attempt to get all correct School names from EMIS to save trouble of further
    # building the hard coded schools_lookup_from_exams_byname
    df['SchoolNameFinal'] = df['SchoolIDFinal'].map(schools_lookup)

    # If does not have NaNs 
    #if not df['SchoolNameEMIS'].hasnans:
    #    # then we can use SchoolNameEMIS as the new final SchoolName
    #    df['SchoolNameFinal'] = df['SchoolIDFinal'].map(schools_lookup)
    #else:
    #    # Unfortunately need to fix this by hand
    #    df['SchoolNameFinal'] = df['SchoolIDFinal'].map(schools_lookup)

    # Check if there is a school that does not have a known
    # mapping either from the EMIS' df_schools or the manually
    # maintained above mapping (old ID, incorrect ones, etc.)
    # If True look at the source file
    if df['SchoolNameFinal'].isnull().values.any():
        print('SchoolID/SchoolName still unknown/check source excel file {}'.format(name)) #df[:1].iloc[:, : 10].to_csv(index=False, header=False)        
        
        # How to clean the data based on configuration (fix in source data or using hand mapping)
        if fix_schoolid_in_source_data:
            df_school_name_missing = df[df['SchoolNameFinal'].isna()]
            print("You have a bad school ID ({}) in your file {}".format(df_school_name_missing['SchoolID'].unique(), name))
        else:
            # Unfortunately need to fix this by hand (mostly in RMI)
            print('All school name and island name combination not yet part of hard coded mapping (if none listed, they likely have a mapping but school is not yet in EMIS):')
            #print('SchoolNameFinal DataFrame with missing values:')
            #display(df[df['SchoolNameFinal'].isna()])
            unique_combination = set(df['SchoolNameTemp'].unique())
            unique_combination_mapped = set(schools_lookup_from_exams_byname.keys())
            unique_combination_not_mapped = unique_combination.difference(unique_combination_mapped)
            for i in unique_combination_not_mapped:
                print("'" + i + "': '',")
    if testing:
        print("DataFrame with records with schoolID still unknown in excel file {}.".format(name))
        display(df[df['SchoolNameFinal'].isnull()])
    
    df = df.drop(labels = ['SchoolID','SchoolName','SchoolNameTemp','SchoolIDTemp1','SchoolIDTemp2'], axis=1)
    df = df.rename(columns = {'SchoolIDFinal': 'SchoolID','SchoolNameFinal': 'SchoolName'})

    if testing:
        print('Cleaned schools DataFrame from file {}.'.format(name))
        display(df)
        
    return df
        
def clean_items(df, name, testing=False):
    """ Does any cleanup/validation needed with Items (test responses.)

    Parameters
    ----------
    df : DataFrame, required
        The student results and enrol DataFrame with items to clean    
    name: str, required
        The name of the excel file this DataFrame came from
    testing: bool, required
        Whether we are test (usually single DataFrame) 
        
    Raises
    ------
    NotImplementedError
        Could raise unknown error. Implement if it happens
    
    Returns
    -------
    DataFrame
    """ 

    # We need to know what test we are cleaning items for since 
    # is affects the validation
    test = df['TestID'].iloc[1]
    
    def strip_spaces(i):
        if isinstance(i,str):
            return i.strip()
        else:
            return i
    
    def validate_answer(i):
        """
        Description:
        validate if the item answer is correct (i.e. It is either 'A', 'B', 'C', 'D' or 'BLANK')
        For high school test the last item column can also be a numeric.

        Args:
            i (Object): item to be validated

        Returns:
            True if item is valid
            False if false is not valid
        """
        if i == 'A' or i == 'B' or i == 'C' or i == 'D' or i == 'BLANK' or i == 'MULT':            
            return True
        elif test == 'H08': #and isinstance(i, float):
            # Numbers can come in as string, in particular when the
            # Series contains some error (i.e. 1` instead of 1)
            # So we test for numeric by attempting a cast
            try:
                float(i)
                return True
            except:
                return False
        else:
            if testing:
                print('Bad Item answer {}'.format(i))
            return False

    def simplify_items(x):
        """ If column an item make it uppercase and strip the redundant string
        """
        if x.startswith('Item_'):  
            if remove_items_metadata:
                item = '_'.join(x.split('_', 2)[:2])
            else:
                item = '_'.join(x.split('_', 2))
            return str.upper(item)
        return x 
          
    # Re-arrange and rename item columns
    df = df.rename(columns = simplify_items)
    
    # Get list of items columns
    cols = df.columns[df.columns.str.startswith('ITEM_')].tolist()
    if testing: 
        print('Cleaned items columns:', cols)
        print('Cleaned items columns length:', len(cols))        
        print('Cleaned items Item-only DataFrame.columns length:', len(df[cols].columns))
        print('Cleaned items Item-only DataFrame.columns')
        display(df[cols].columns)
    
    # Strip out any spaces from all ITEMS
    try:
        df[cols] = df[cols].map(strip_spaces)
    except:
        print('Detected a mismatch in item numbers (e.g. repeating Item_039, Item_039, etc.) in excel file {}'.format(name))
    
    try:
        # Remove all multiple answers (all the time?). Answers like (A,C), A&B, etc.
        df[cols] = df[cols].replace(to_replace='\({0,1}[A-D](,|&).*\){0,1}', value='MULT', regex=True)
    
        # Insert string 'BLANK' where na
        df[cols] = df[cols].fillna('BLANK')
        
        # Insert string 'BLANK' where there is whitespace character(s)
        df[cols] = df[cols].replace(to_replace='^\s+&', value='BLANK', regex=True)
        
        # Insert string 'BLANK' where there is an empty string like ''
        df[cols] = df[cols].replace(to_replace='', value='BLANK', regex=False)
    except ValueError as e:
        cols1 = len(cols)
        cols2 = len(df[cols].columns)
        print('Cleaned items possible duplicate item. Columns starting with ITEM_ is {} while DataFrame columsn is {} (hint from data): '.format(cols1, cols2),
              df[:1].iloc[:, : 10].to_csv(index=False, header=False))
        print('Error was: ', e)
    
    df[cols].apply(lambda x: x.astype(str).str.upper())
    
    # Finally, validate the answer for all the items.
    #cols2 = list(cols)    

    if test == 'H08':
        h08_cols = list(cols)
        # High School so need to do two runs
        # One for the final item which has numeric values
        last_col = h08_cols.pop()            
        df_last_col_mask = df[[last_col]].map(validate_answer)

        # Two for all standard multiple choice questions    
        df_standard_mask = df[h08_cols].map(validate_answer)
        
        if testing:
            print("last_col: ", last_col)            
            print("h08_cols: ", h08_cols)
            display(df[[last_col]].map(validate_answer))
            display(df[[last_col]].map(validate_answer).all().all())
            display(df[h08_cols].map(validate_answer))
            display(df[h08_cols].map(validate_answer).all().all())
        
        if not (df_last_col_mask.all().all() and df_standard_mask.all().all()):
            
            # Collect invalid items to return very specific feedback
            invalid_items = []
            for c in h08_cols: # errors in standard columns
                if len(df_standard_mask[~df_standard_mask[c]].index.values) != 0:
                    invalid_items.append(c+" rows: "+str(df_standard_mask[~df_standard_mask[c]].index.values + 2))
                    
            if len(df_last_col_mask[~df_last_col_mask[last_col]].index.values) != 0: # errors in last columns
                invalid_items.append(last_col+" rows: "+str(df_last_col_mask[~df_last_col_mask[last_col]].index.values + 2))
            
            # Do not even flag incorrect answers
            if not skip_incorrect_answers:
                print('Invalid answers detected in test {} (note: supposedly a High School test) year {} (from excel file {}). Invalid answers are in {}'.format(
                    df['TestName'].iloc[0], df['SchoolYear'].iloc[0], name, str(invalid_items)))
            elif skip_incorrect_answers:
                # but clean them up instead
                df[cols] = df[cols].where(df_standard_mask, 'BLANK')

    else:
        # Any other test (i.e. not High School)
        df_standard_mask = df[cols].map(validate_answer)
        
        if testing:
            print('Cols {}'.format(cols))
        
        if testing:
            print('Items errors mask DataFrame:')            
            display(df[df_standard_mask.eq(False).any(axis=1)])
            
        if not df_standard_mask.all().all():
            
            # Collect invalid items to return very specific feedback
            invalid_items = []
            for c in cols: # errors in all columns
                if len(df_standard_mask[~df_standard_mask[c]].index.values) != 0:
                    invalid_items.append(c+" rows: "+str(df_standard_mask[~df_standard_mask[c]].index.values + 2))
            
            # Do not even flag incorrect answers
            if not skip_incorrect_answers:
                print('Invalid answers detected in test {} year {} (from excel file {}). Invalid answers are in {}'.format(
                    df['TestName'].iloc[0], df['SchoolYear'].iloc[0], name, str(invalid_items)))
            elif skip_incorrect_answers:
                # but clean them up instead
                df[cols] = df[cols].where(df_standard_mask, 'BLANK')
            
    if testing:
        print('Cleaned items DataFrame from file {}.'.format(name))
        display(df)
    
    return df

def clean_students(df, name, testing=False):
    """ Cleanup students data here. There is stuff from other functions like in the merge
    above that could be put more cleanly here.

    Parameters
    ----------
    df : DataFrame, required
        The student results and enrol DataFrame
    name: str, required
        The name of the excel file this DataFrame came from
    testing: bool, required
        Whether we are test (usually single DataFrame) 
        
    Raises
    ------
    NotImplementedError
        Could raise unknown error. Implement if it happens
    
    Returns
    -------
    DataFrame
    """
    
    genders = {
        'Male': 'M',
        'MALE': 'M',
        'm': 'M',       
        'M': 'M',       
        'M,': 'M',
        'M ': 'M',
        'Boy': 'M',
        'BOY': 'M',
        'b': 'M',        
        'B': 'M',
        'n': 'M',
        'mm': 'M',
        'MM': 'M',
        'FM': 'M',        
        'm  m': 'M',
        'Bm': 'M',
        'Female': 'F',        
        'FEMALE': 'F',
        'f': 'F',       
        'F': 'F',
        'Girl': 'F',        
        'GIRL': 'F',
        'g': 'F',        
        'G': 'F',
        #'I': 'F',
        '`f': 'F',
        'FF': 'F',
        'ff': 'F',
        #'?': 'M',
        #'??': 'M',
        #'???': 'M',
        #'????': 'M',
        #'BLANK': 'M',
        #'a': 'F',
        #'A': 'F',
        'BLANKm': 'M',
        #',': 'M',
        #'ERROR #3100': 'M',
        #'**': 'M',
    }
    
    yesno = {
        'Yes': 'Yes',        
        'YEs': 'Yes',
        'YES': 'Yes',        
        'yes': 'Yes',
        'y': 'No',
        'No': 'No',        
        'NO': 'No',
        'no': 'No',        
        'n': 'No',
    }
    
    if testing:
        print('Cleaning students...')
        
    # Strip any extra spaces for student names
    try:
        df['StudentName'] = df['StudentName'].apply(lambda x: x.strip())
    except Exception as e:
        print('Student name with non-string data detected in test {} year {} (from excel file {})'.format(
            df['TestName'].iloc[0], df['SchoolYear'].iloc[0], name))
        
    # Adjust the correct student ID where possible and generate UUID for all others
    
    # At this point student ID already there come from the EMIS
    # Perhaps it might be useful to build a list of automatically assigned
    # UUIDs as they get processed and look there as well?!
    missing_student_id_tot = df['stuCardID'].isna().sum()
    missing_student_ids = []

    for i in range(missing_student_id_tot):
        missing_student_ids.append(uuid.UUID(int=rd.getrandbits(128), version=4))

    df.loc[df.stuCardID.isnull(), 'stuCardID'] = missing_student_ids    
        
    # Coalesce student genders
    df['GenderFinal'] = df.stuGender.combine_first(df.Gender)
    
    # Clean genders (should we also flag?)
    df['GenderFinal'] = df['GenderFinal'].str.strip() # strip out leading/trailing spaces
    df['GenderFinal'] = df['GenderFinal'].map(genders)
    # Tell me if the DataFrame has any unkown gender
    df_badgenders = df_onlinesba[~df_onlinesba['Gender'].isin(list(genders.keys()))]
    
    if not df_badgenders.empty:
        if accept_unknown_gender:
            # Accept but "fix" the bad gender (is there anything to fix?)
            pass
        else:
            print('Some unknown gender detected (these ones: {}) detected in test {} year {} (from excel file {})'.format(
                set(df_badgenders['Gender'].unique()), df['TestName'].iloc[0], df['SchoolYear'].iloc[0], name))
        if testing:
            print('These records have bad genders')
            display(df[df['GenderFinal'].isna()])
    
    # Clean some boolean (should we flag?)
    df['SpEdCode'] = df['SpEdCode'].map(yesno)
    df['SpEdCode'] = df['SpEdCode'].fillna('No')        
    df['Accommodation'] = df['Accommodation'].map(yesno)
    df['Accommodation'] = df['Accommodation'].fillna('No')
    
    # Student names with * or ??? to be flagged and/or handled
    if df['StudentName'].str.contains('^ *\?+ *$', na=False, regex=True).any() and not accept_unknown_student:
        print('Student name with ??? (unknown/bad student name) detected in test {} year {} (from excel file {})'.format(
            df['TestName'].iloc[0], df['SchoolYear'].iloc[0], name))
        
    df['StudentName'] = df['StudentName'].replace(to_replace=r'^ *\?+ *$', value=np.nan, regex=True)
    
    # Student NaN
    if df['StudentName'].isnull().values.any() and not accept_unknown_student:
        print('Student name without any value (no student name) detected in test {} year {} (from excel file {})'.format(
            df['TestName'].iloc[0], df['SchoolYear'].iloc[0], name))
        if testing:
            display(df.loc[df['StudentName'].isnull()])
    
    # Replace all unknown with 'Unknown student 1', 'Unknown student 2', etc.    
    #df['StudentName'] = df['StudentName'].fillna('Unknown')
    mask = df['StudentName'].isna()
    values = [f'Unknown student {x}' for x in range(1, mask.sum()+1)]        
    df.loc[mask, 'StudentName'] = values
    
    # Student names repeating
    #df.duplicated(subset=['StudentName'], keep='first')
    df_dups = df.duplicated(subset=['StudentName'], keep='first')
        
    if flag_duplicate_students:
        print('Students that looks like duplicates:')
        df = df[df.duplicated(subset=['StudentName'], keep='first')]    
    else:
        # Clean duplicates by adding ? as middle name
        if testing:
            print('Cleaning duplicate students')
        df['StudentName'] = df['StudentName'].where(~df.duplicated(subset=['StudentName'], keep='first'), df['StudentName'].add(' ?'))
    
    if testing:
        print('Cleaned students DataFrame from file {}.'.format(name))
        display(df)
    
    df = df.drop(labels = ['Gender'], axis=1)
    df = df.rename(columns = {'GenderFinal': 'Gender'})
        
    return df

def clean_teachers(df, name, testing=False):
    """ Mostly a placeholder at the moment. But might be desirable to cleanup
    teachers, add teacher ID, etc.

    Parameters
    ----------
    df : DataFrame, required
        The student results (contains teachers) and enrol DataFrame
    name: str, required
        The name of the excel file this DataFrame came from
    testing: bool, required
        Whether we are test (usually single DataFrame) 
        
    Raises
    ------
    NotImplementedError
        Could raise unknown error. Implement if it happens
    
    Returns
    -------
    DataFrame
    """
    
    # Teachers containing numbers instead of names
    if pd.to_numeric(df_onlinesba['Teacher'], errors='coerce').any():
        print('Teacher with numeric values detected in test {} year {} (from excel file {})'.format(
            df['TestName'].iloc[0], df['SchoolYear'].iloc[0], name))
    # Clean it. The cleaning here is temporary and only so the remaining code works
    # this should really cleaned up in source data
    df['Teacher'] = df['Teacher'].astype('string')
    
    # Fill missing teachers
    df['Teacher'] = df['Teacher'].fillna('N/A')
    
    # Teachers names with ??? (Flag it)
    if df['Teacher'].str.contains(' *\?+ *', na=False, regex=True).any() and not accept_unknown_teacher:
        print('Teacher with ?? as names detected in test {} year {} (from excel file {})'.format(
            df['TestName'].iloc[0], df['SchoolYear'].iloc[0], name))
    # Clean it
    df['Teacher'] = df['Teacher'].replace(to_replace=' *\?+ *', value='Unknown', regex=True)
    
    # Teachers names with less then 3 alphanumeric charactors (Flag it)
    if df['Teacher'].str.contains('^\w{1,3}$', na=False, regex=True).any() and not accept_teachers_with_three_chars_only:        
        print('Teacher names with 3 characters or less detected (valid name?) in test {} year {} (from excel file {})'.format(
            df['TestName'].iloc[0], df['SchoolYear'].iloc[0], name))
    # Clean it ?
    #df['Teacher'] = df['Teacher'].replace(to_replace='\w{1,3}', value='Unknown', regex=True)
    
    # Teachers without names
    if df['Teacher'].str.contains('^\s+$', na=False, regex=True).any() and df['Teacher'].hasnans and not accept_unknown_teacher:
        print('Teacher without names detected in test {} year {} (from excel file {})'.format(
            df['TestName'].iloc[0], df['SchoolYear'].iloc[0], name))
    # Clean it
    df['Teacher'] = df['Teacher'].replace(to_replace='^\s+$', value='Unknown', regex=True)
    
    if testing:
        print('Cleaned teachers DataFrame from file'.format(name))
        display(df)
        
    return df

def convert_to_onlinesba(df, name, testing=False):
    """ A pretty fat function that does some validation, cleaning and converting
    to the OnlineSBA format. Function can be split if needed as this tools gets refine
    through practicalities of real life usage

    Parameters
    ----------
    df : DataFrame, required
        The student results and enrol DataFrame
    name: str, required
        The name of the excel file this DataFrame came from
    testing: bool, required
        Whether we are test (usually single DataFrame) 
        
    Raises
    ------
    NotImplementedError
        Could raise unknown error. Implement if it happens
    
    Returns
    -------
    DataFrame
    """
    
    df = df.rename(columns = {
        'stuCardID': 'STUDENTID', 
        'SpEdCode': 'SPED',
        'Accommodation': 'ACCOM',
        'StudentName': 'STUDENTNAME',
        'SchoolID': 'SCHOOLID',
        'Gender': 'GENDER',
        'TestID': 'TESTID',
        'Teacher': 'TEACHERNAME',
        'SchoolYear': 'SCHOOLYEAR'
        })
    df = df.drop(labels = [
        'RecordNo', 'TestName', 'IslandName', # 'SchoolYear',
        'SchoolName', 'StudentID', 'Ethnicity', 'Disability', 
        'ELL', 'Migrant', 'FRLunch', 'StudentName2', 'Student',
        'stuGender', 'stuDoB', 'schNo', 'stueYear', 'Student2'], axis=1, errors='ignore')

    cols = list(df.columns)
    cols_items = [i for i in cols if 'ITEM_' in i]
    cols_items.sort()

    # Re-order. First set of colums will likely always be there and the same
    # followed by a varying number of exam items
    df = df[[
        'SCHOOLYEAR','STUDENTID','SPED','ACCOM','STUDENTNAME','SCHOOLID','GENDER','TESTID',
        'TEACHERNAME'] + cols_items]
    
    if testing:
        print('Final OnlineSBA DataFrame for data from file {}.'.format(name))
        display(df)
    
    return df


# %%
# Cleanup DataFrame as OnlineSBA input format
# Working with the single student exams file (for testing)
df_onlinesba = clean_exams(df_students_results_and_enrol[testname], testname, testing=True)
df_onlinesba = clean_schools(df_onlinesba, df_schools, testname, testing=True)
df_onlinesba = clean_items(df_onlinesba, testname, testing=True)
df_onlinesba = clean_students(df_onlinesba, testname, testing=True)
df_onlinesba = clean_teachers(df_onlinesba, testname, testing=True)
df_onlinesba = convert_to_onlinesba(df_onlinesba, testname, testing=True)

# %%
# %%time
# Cleanup DataFrame as OnlineSBA input format
# Working with all student exams files (~17 seconds on iMac with i9 CPU and 32GB RAM)
df_onlinesba_dict = {}

print('Processing exams data from excel files...')

for file, df in tqdm(df_students_results_and_enrol_list.items()):
    #tqdm.write('Processing exam ID {} for year {} from excel file {}'.format(df['TestID'].values[0], df['SchoolYear'].values[0], file))
    df_onlinesba = clean_exams(df_students_results_and_enrol_list[file], file, testing=False)
    df_onlinesba = clean_schools(df_onlinesba, df_schools, file, testing=False)
    df_onlinesba = clean_items(df_onlinesba, file, testing=False)
    df_onlinesba = clean_students(df_onlinesba, file, testing=False)
    df_onlinesba = clean_teachers(df_onlinesba, file, testing=False)
    df_onlinesba = convert_to_onlinesba(df_onlinesba, file, testing=False)
    df_onlinesba_dict[file] = df_onlinesba

len(df_onlinesba_dict)

# %%
# Write processed data back into excel (or CSV directly)
# Working with the single student exams file (for testing)
filename_xls = os.path.join(local_path,country+'/onlinesba-test.xlsx')
filename_csv = os.path.join(local_path,country+'/onlinesba-test.csv')

schoolyear = df_onlinesba['SCHOOLYEAR'].values[0]
testid = df_onlinesba['TESTID'].values[0]

with pd.ExcelWriter(filename_xls) as writer:
    # add DataFrames you want to write to Excel here
    df_onlinesba.to_excel(writer, index=False, sheet_name='Sheet1', engine='openpyxl')
    wb = writer.book
    ws = wb.create_sheet(title='ExamYear')
    ws['A1'] = '20'+schoolyear.split('-')[1]

df_onlinesba.to_csv(filename_csv, index=False)

# %%
# Testing
display(list(df_onlinesba_dict.keys()))
display(df_onlinesba_dict[list(df_onlinesba_dict.keys())[0]])
display(df_onlinesba_dict['AllSchools_A03_2008-09_Results.xls'])

# %%
# %%time
# Write processed data back into excel (or CSV directly much faster)
# Working with all student exams files (~1min 52sec on iMac with i9 CPU and 32GB RAM for Excel, 2sec for CSV)

for file, df in tqdm(df_onlinesba_dict.items()):
    schoolyear = df['SCHOOLYEAR'].values[0]
    testid = df['TESTID'].values[0]
    try:
        # Could remove the SCHOOLYEAR if onlinesba really requires it
        if export == 'csv':        
            filename = os.path.join(local_path, country+'/onlinesba-load-files-csv/' + schoolyear + '-' + testid + '.csv')
            df.to_csv(filename, index=False)
        else:
            filename = os.path.join(local_path, country+'/onlinesba-load-files-xls/' + schoolyear + '-' + testid + '.xlsx')
            with pd.ExcelWriter(filename) as writer:
                # add DataFrames you want to write to Excel here
                df.to_excel(writer, index=False, sheet_name='Sheet1', engine='openpyxl')
                # The excel version is used to load into FedEMIS and currently requires
                # a Sheet 'ExamYear' to indicate the year in cell A1                
                #tqdm.write(type(writer))
                wb = writer.book
                ws = wb.create_sheet(title='ExamYear')
                ws['A1'] = '20'+schoolyear.split('-')[1]
                
        #tqdm.write('Writing file {}'.format(filename))
    except TypeError:
        tqdm.write('Problem with a type, cannot generate filename {}'.format(filename))
    except Exception as e:
        tqdm.write('Unknown error {} with file {}'.format(e, filename))

# %%
# Get the exact matches (i.e. exact name in exams data and the EMIS)
# Working with the single student exams file (for testing)
df_students_results_and_enrol_example = df_students_results_and_enrol[list(df_students_results_and_enrol.keys())[0]]
df_exact_matches = df_students_results_and_enrol_example.dropna(how='all', subset=['stuCardID']) #subset=['stuCardID', 'stuGender', 'stuDoB', 'schNo', 'stueYear'])
display(df_exact_matches)

# %%
# %%time
# Get the exact matches (i.e. exact name in exams data and the EMIS)
# Working with all student exams files (~23 seconds on iMac with i9 CPU and 32GB RAM)
df_exact_matches_list = []
for file, df in df_students_results_and_enrol_list.items():
    df_exact_matches_list.append(df.dropna(how='all', subset=['stuCardID']))

# %%
# #%%time
# WARNING: Not currently running as df_student_enrol_nonambiguous is no longer globally defined
# Commenting this out if ever needed

# Just included for playing around. Not currently being used, just working with exact matches for now

# Here we will get a bit more sophisticated in trying to match students to get their EMIS
# canonical data (DoB, ID, etc.)
# Instead of doing a simple name matching we will do fuzy search using the Levenshtein algorithm
# That way we will capture students with slightly different name spellings

# Is this time consuming search worth it?!

#import fuzzy_pandas as fpd

#exams_cols = list(set(df_student_results.columns))
#stuen_cols = list(set(df_student_enrol_nonambiguous.columns))

# the threshold is set high so we may not capture students with terribly
# bad spellings but will capture things with only small mis-spelling
# and reduce chances of false positive matching
#df_fuzzy_matches = fpd.fuzzy_merge(
#    df_student_results, df_student_enrol_nonambiguous,
#    left_on=['StudentName2'], right_on=['Student2'],
#    #keep='all',
#    method='levenshtein',
#    threshold=0.94, #0.9
#    ignore_case=True,
#    ignore_nonalpha=False,
#    ignore_nonlatin=False,
#    ignore_order_words=False,
#    ignore_order_letters=False,
#    ignore_titles=False,
#    join='left-outer' # { 'inner', 'left-outer', 'right-outer', 'full-outer' }
#)

#df_fuzzy_matches

#s = df_fuzzy_matches['stuCardID'] == ''
#s.sum()

# %%
# Write various DataFrame into Excel to examine (testing)
filename = os.path.join(local_path, country+'/soe-to-online-test.xlsx')

df_student_results_example = df_student_results[list(df_student_results.keys())[0]]
df_students_results_and_enrol_example = df_students_results_and_enrol[list(df_students_results_and_enrol.keys())[0]]

with pd.ExcelWriter(filename) as writer:
    # add DataFrames you want to write to Excel here
    df_student_results_example.to_excel(writer, index=False, sheet_name='Sheet1', engine='openpyxl')
    df_students_results_and_enrol_example.to_excel(writer, index=False, sheet_name='Sheet2', engine='openpyxl')
    #df_fuzzy_matches.to_excel(writer, index=False, sheet_name='Sheet3', engine='openpyxl')
    df_onlinesba.to_excel(writer, index=False, sheet_name='Sheet4', engine='openpyxl')

# %%
